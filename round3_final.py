"""
Round 3 Final Update for 2026 IPOs
====================================
Updates the last 6 TODO 2026 stocks (2026/01/02 - 2026/01/12) found via per-stock
web searches (HKET, 信報, etnet, 格隆匯, Yahoo Finance HK, HSTONG, Eastmoney).

BONUS: Grey market data also collected opportunistically for 4 of them.
"""
from openpyxl import load_workbook

INPUT  = "HK_IPO_Full_Analysis_Report_v2_ROUND2.xlsx"
OUTPUT = "HK_IPO_Full_Analysis_Report_v2_FINAL.xlsx"
SHEET = "IPO全量數據分析"
PLACEHOLDER = "待公告/需手動核查"

COL = {
    "Stock Code":1,
    "公開零售認購倍數":6, "國際配售認購倍數":7,
    "暗盤收盤回報率 (%)":11, "首日收市回報率 (%)":12, "一手中簽率 (%)":14,
}

# Data per stock code. None = leave as 待公告, skip.
# Sources cited in the comment next to each value.
DATA = {
    "06082": {  # 壁仞科技 2026/01/02
        "公開零售認購倍數":   "2347.53倍",    # 格隆匯: 公開發售2347.53倍
        "一手中簽率 (%)":     "5.0%",          # HKET: 一手中籤率5%
        "暗盤收盤回報率 (%)": "+79.69%",       # 格隆匯: 暗盤收升79.69%
        "首日收市回報率 (%)": "+75.82%",       # 格隆匯: 首日收盤較發行價大升75.82%
    },
    "02513": {  # 智譜 2026/01/08
        "公開零售認購倍數":   "1158.46倍",    # 信報: 公開發售1158.46倍
        "一手中簽率 (%)":     "5.0%",          # 信報: 中籤率5%
        "首日收市回報率 (%)": "+13.17%",       # 東方財富: 131.50/116.20-1 = +13.17%
    },
    "09903": {  # 天數智芯 2026/01/08
        "公開零售認購倍數":   "413.24倍",     # 信報: 公開發售413.24倍
        "一手中簽率 (%)":     "7.0%",          # 信報: 中籤率7%
        "首日收市回報率 (%)": "+8.44%",        # 阿思達克: 156.8/144.6-1 = +8.44%
    },
    "02675": {  # 精鋒醫療-B 2026/01/08
        "公開零售認購倍數":   "1091.94倍",    # 東方財富: 公配1091.94倍
        "一手中簽率 (%)":     "0.5%",          # 信報: 中籤率0.5%
        "暗盤收盤回報率 (%)": "+37.84%",       # 東方財富(輝立): 59.60 vs 43.24
        "首日收市回報率 (%)": "+30.90%",       # 東方財富: 56.60/43.24-1
    },
    "06938": {  # 瑞博生物-B 2026/01/09
        # 公開認購最終倍數未查到明確最終數字,略過
        "首日收市回報率 (%)": "+41.63%",       # etnet+HSTONG: 82.10/57.97-1
    },
    "00501": {  # 豪威集團 2026/01/12
        "公開零售認購倍數":   "9.28倍",       # 匯港通訊: 公開發售8.28倍; 阿思達克 9.28倍(差異,取後者)
        "國際配售認購倍數":   "9.73倍",       # 阿思達克: 國際發售9.73倍(額外補欄位7)
        "一手中簽率 (%)":     "100.0%",        # 匯港/阿思達克: 一手中籤率100%
        "暗盤收盤回報率 (%)": "-0.30%",         # 阿思達克(富途): 104.5/104.8-1
        # 首日收市百分比未查到明確數字,略過
    },
}


def main():
    wb = load_workbook(INPUT)
    ws = wb[SHEET]
    code_to_row = {str(ws.cell(row=r, column=1).value).strip(): r
                   for r in range(2, ws.max_row+1) if ws.cell(row=r, column=1).value}

    cells_changed = 0
    touched_codes = []
    log = []
    for code, fields in DATA.items():
        if code not in code_to_row:
            log.append(f"✗ {code} not in user file"); continue
        row = code_to_row[code]
        changed_this = 0
        for field, new_v in fields.items():
            col = COL[field]
            old_v = ws.cell(row=row, column=col).value
            if old_v != PLACEHOLDER:
                continue
            ws.cell(row=row, column=col).value = new_v
            cells_changed += 1
            changed_this += 1
        if changed_this:
            touched_codes.append(code)
            log.append(f"✓ {code}: {changed_this} cells updated")

    wb.save(OUTPUT)
    print("\n".join(log))
    print(f"\nTotal cells updated this round: {cells_changed}")
    print(f"Stocks touched: {len(touched_codes)}")

    # Final coverage analysis
    print("\n" + "=" * 60)
    print("FINAL COVERAGE REPORT")
    print("=" * 60)
    stats = {
        "公開零售認購倍數": 0, "國際配售認購倍數": 0, "國際配售認購人數": 0,
        "基石投資者占比 (%)": 0, "公司大股東持股占比 (上市後 %)": 0,
        "暗盤收盤回報率 (%)": 0, "首日收市回報率 (%)": 0,
        "一個月回報率 (%)": 0, "一手中簽率 (%)": 0,
    }
    col_map = {"公開零售認購倍數":6, "國際配售認購倍數":7, "國際配售認購人數":8,
               "基石投資者占比 (%)":9, "公司大股東持股占比 (上市後 %)":10,
               "暗盤收盤回報率 (%)":11, "首日收市回報率 (%)":12,
               "一個月回報率 (%)":13, "一手中簽率 (%)":14}
    total = ws.max_row - 1
    for r in range(2, ws.max_row+1):
        for field, col in col_map.items():
            if ws.cell(row=r, column=col).value != PLACEHOLDER:
                stats[field] += 1
    print(f"\n{'Field':<35}{'Filled':<10}{'Coverage'}")
    print("-"*60)
    for f, c in stats.items():
        print(f"{f:<35}{c}/{total:<8}{100*c/total:.1f}%")

if __name__ == "__main__":
    main()
